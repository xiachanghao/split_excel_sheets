import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment, Protection, GradientFill
from openpyxl.styles.colors import Color
from copy import copy
from openpyxl.utils import get_column_letter
import sys

def copy_styles(source_cell, target_cell):
    if source_cell.has_style:
        target_cell.font = Font(**source_cell.font.__dict__)
        target_cell.border = Border(
            left=Side(**{k: v for k, v in source_cell.border.left.__dict__.items() if k != '_parent'}),
            right=Side(**{k: v for k, v in source_cell.border.right.__dict__.items() if k != '_parent'}),
            top=Side(**{k: v for k, v in source_cell.border.top.__dict__.items() if k != '_parent'}),
            bottom=Side(**{k: v for k, v in source_cell.border.bottom.__dict__.items() if k != '_parent'})
        )
        target_cell.fill = copy_fill(source_cell.fill)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = Protection(**source_cell.protection.__dict__)
        target_cell.alignment = Alignment(**source_cell.alignment.__dict__)

def copy_fill(fill):
    # 处理弃用警告
    if hasattr(fill, '__dict__'):
        fill = copy(fill)
    if isinstance(fill, PatternFill):
        return PatternFill(
            patternType=fill.patternType,
            fgColor=copy_color(fill.fgColor),
            bgColor=copy_color(fill.bgColor),
            start_color=copy_color(fill.start_color),
            end_color=copy_color(fill.end_color)
        )
    elif isinstance(fill, GradientFill):
        stops = [(stop.position, copy_color(stop.color)) for stop in fill.stops]
        return GradientFill(
            degree=fill.degree,
            stops=stops,
            type=fill.type,
            gradient_type=fill.gradient_type
        )
    return fill

def copy_color(color):
    if color is None:
        return None
    return Color(rgb=color.rgb)

def copy_row_heights(source_sheet, target_sheet):
    for row in range(1, source_sheet.max_row + 1):
        target_sheet.row_dimensions[row].height = source_sheet.row_dimensions[row].height

def copy_column_widths(source_sheet, target_sheet):
    for col in range(1, source_sheet.max_column + 1):
        # 处理属性错误
        column_letter = get_column_letter(col)
        target_sheet.column_dimensions[column_letter].width = \
            source_sheet.column_dimensions[column_letter].width

def copy_merged_cells(source_sheet, target_sheet):
    for merged_range in source_sheet.merged_cells.ranges:
        target_sheet.merge_cells(str(merged_range))

def split_excel_sheets(input_file):
    # 加载整个Excel文件
    wb = load_workbook(input_file, keep_vba=True)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # 创建一个新的工作簿和工作表
        new_wb = Workbook()
        new_ws = new_wb.active
        new_ws.title = sheet_name
        
        # 将数据写入新的工作表
        for r_idx, row in enumerate(ws.rows, start=1):
            for c_idx, cell in enumerate(row, start=1):
                new_cell = new_ws.cell(row=r_idx, column=c_idx, value=cell.value)
                copy_styles(cell, new_cell)
        
        # 复制行高
        copy_row_heights(ws, new_ws)
        
        # 复制列宽
        copy_column_widths(ws, new_ws)
        
        # 复制合并单元格
        copy_merged_cells(ws, new_ws)
        
        # 保存新的工作簿
        new_wb.save(f"{sheet_name}.xlsx")
        print(f"Saved {sheet_name}.xlsx")

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python script.py <path_to_excel_file>")
        sys.exit(1)
    
    input_file = sys.argv[1]
    split_excel_sheets(input_file)